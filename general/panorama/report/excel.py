"""Excel (.xlsx) report with sheets: FRONTPAGE, SAST, SCA, INFRA, LICENSES.

Supports an optional Markdown-based XLSX template to define table-like sheets
using paths against the canonical JSON report (sast.findings,
dependency_vulnerabilities.vulnerabilities, sbom.components, etc.). If a
template is configured and can be parsed, SAST/SCA/LICENSES table sheets are
generated from it; FRONTPAGE and INFRA keep using the existing code layout.
"""
from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from analyze.filters import severity_for_vuln


log = logging.getLogger(__name__)


def _safe(v: Any) -> str:
    return "" if v is None else str(v)


def _sca_severity_display(sev: str) -> str:
    """Show only severity level (e.g. HIGH), stripping any leading numeric prefix like '0 '."""
    if not sev or not isinstance(sev, str):
        return _safe(sev)
    s = str(sev).strip()
    parts = s.split(None, 1)
    if len(parts) == 2 and parts[0].isdigit():
        return parts[1]
    return s


def _severity_sast(findings: list[dict]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for f in findings:
        rid = f.get("rule_id") or ""
        if rid.startswith("deps.") or rid.startswith("infra."):
            continue
        sev = _safe(f.get("severity", "unknown")).upper()
        counts[sev] = counts.get(sev, 0) + 1
    return counts


def _severity_sca(vulns: list[dict]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for v in vulns:
        raw = v.get("severity") or severity_for_vuln(_safe(v.get("vuln_id", "")))
        sev = _sca_severity_display(raw)
        counts[sev] = counts.get(sev, 0) + 1
    return counts


def _purl(eco: str, name: str, version: str) -> str:
    eco_lower = (eco or "").lower().replace(" ", "")
    return f"pkg:{eco_lower}/{name}@{version}"


def _get_path(obj: Any, path: str) -> Any:
    """Resolve a dotted path like 'sast.findings' or 'id' against nested dicts."""
    if not path:
        return obj
    parts = [p for p in str(path).split(".") if p]
    cur: Any = obj
    for part in parts:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(part)
        if cur is None:
            return None
    return cur


def _parse_bool(value: str | bool | None) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if not s:
        return None
    if s in ("1", "true", "yes", "on"):
        return True
    if s in ("0", "false", "no", "off"):
        return False
    return None


def _parse_xlsx_template(template_path: str) -> dict[str, Any]:
    """Parse Markdown XLSX template into sheets and tables.

    Soporta:

    - Una tabla por hoja usando el formato clásico:

      ## Sheet SAST (source: sast.findings)

      | Header | Value path | Width | Wrap |
      |--------|----------- |-------|------|
      | ...    | ...        | ...   | ...  |

    - Varias tablas por hoja usando bloques con metadatos ligeros:

      ## Sheet INFRA

      Source: infrastructure.images
      | Header | Value path | Width | Wrap |
      |--------|----------- |-------|------|
      | ...    | ...        | ...   | ...  |

      Source: infrastructure.findings
      | Header | Value path | Width | Wrap |
      |--------|----------- |-------|------|
      | ...    | ...        | ...   | ...  |

      Source: infrastructure.findings
      Expand: vulnerabilities
      | Header | Value path | Width | Wrap |
      |--------|----------- |-------|------|
      | ...    | ...        | ...   | ...  |

    Donde:
    - `Source:` define la ruta en el JSON canónico para esa tabla concreta.
    - `Expand: vulnerabilities` indica que se debe aplanar la lista
      `vulnerabilities` de cada entrada (útil para CVEs de INFRA).
    """
    with open(template_path, "r", encoding="utf-8") as fh:
        lines = fh.readlines()

    sheets: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    current_source: str | None = None
    current_expand: str | None = None

    def finish_current() -> None:
        nonlocal current
        nonlocal current_source, current_expand
        if current is not None:
            tables = current.get("tables") or []
            # Compatibilidad: si no se han detectado tablas por bloque pero hay columnas
            # antiguas, crear una tabla única.
            if not tables and current.get("columns"):
                tables = [
                    {
                        "source": current.get("source"),
                        "columns": current["columns"],
                    }
                ]
                current["tables"] = tables
            if current.get("name") and current.get("type", "table") == "table" and tables:
                sheets.append(current)
        current = None
        current_source = None
        current_expand = None

    i = 0
    n = len(lines)
    while i < n:
        line = lines[i].rstrip("\n")
        if line.startswith("## "):
            # New sheet header?
            head = line[3:].strip()
            if head.lower().startswith("sheet "):
                finish_current()
                rest = head[len("sheet ") :].strip()
                name = rest
                attrs: dict[str, str] = {}
                if "(" in rest and rest.endswith(")"):
                    name_part, attrs_part = rest.split("(", 1)
                    name = name_part.strip()
                    attrs_str = attrs_part[:-1]
                    for part in attrs_str.split(","):
                        if ":" in part:
                            k, v = part.split(":", 1)
                            attrs[k.strip().lower()] = v.strip()
                current = {
                    "name": name,
                    "type": attrs.get("type", "table").lower(),
                    "source": attrs.get("source"),
                    "columns": [],
                    "tables": [],
                }
            i += 1
            continue

        if current is not None:
            stripped = line.strip()
            # Metadatos por tabla dentro de una hoja
            if stripped.lower().startswith("source:"):
                current_source = stripped.split(":", 1)[1].strip()
                i += 1
                continue
            if stripped.lower().startswith("expand:"):
                current_expand = stripped.split(":", 1)[1].strip().lower()
                i += 1
                continue

            # Tabla Markdown
            if line.lstrip().startswith("|"):
                table_lines: list[str] = []
                while i < n and lines[i].lstrip().startswith("|"):
                    table_lines.append(lines[i].rstrip("\n"))
                    i += 1
                if len(table_lines) >= 2:
                    header_line = table_lines[0]
                    header_cells = [
                        c.strip().lower()
                        for c in header_line.strip().strip("|").split("|")
                    ]

                    def normalize_header(h: str) -> str:
                        h = h.strip().lower()
                        if h in ("header", "column", "title"):
                            return "header"
                        if h in ("value", "value path", "value_path", "path"):
                            return "value"
                        if h == "width":
                            return "width"
                        if h == "wrap":
                            return "wrap"
                        if h in ("number format", "number_format"):
                            return "number_format"
                        if h == "style":
                            return "style"
                        return h

                    norm_headers = [normalize_header(h) for h in header_cells]

                    columns: list[dict[str, Any]] = []
                    # Data rows: skip header and separator
                    for row_line in table_lines[2:]:
                        row_cells = [
                            c.strip()
                            for c in row_line.strip().strip("|").split("|")
                        ]
                        if not any(row_cells):
                            continue
                        row: dict[str, Any] = {}
                        for idx, raw in enumerate(row_cells):
                            if idx >= len(norm_headers):
                                break
                            key = norm_headers[idx]
                            if not key:
                                continue
                            row[key] = raw
                        header_text = row.get("header")
                        value_path = row.get("value")
                        if not header_text or not value_path:
                            continue
                        width_val: Any = row.get("width")
                        width: float | None
                        if width_val is None or width_val == "":
                            width = None
                        else:
                            try:
                                width = float(width_val)
                            except ValueError:
                                width = None
                        wrap = _parse_bool(row.get("wrap"))
                        columns.append(
                            {
                                "header": header_text,
                                "value": value_path,
                                "width": width,
                                "wrap": wrap,
                                "number_format": row.get("number_format"),
                                "style": row.get("style"),
                            }
                        )

                    if columns:
                        table_source = current_source or current.get("source")
                        table: dict[str, Any] = {
                            "source": table_source,
                            "columns": columns,
                        }
                        if current_expand:
                            table["expand"] = current_expand
                        current.setdefault("tables", []).append(table)
                        # Reset metadatos por tabla
                        current_source = None
                        current_expand = None
                continue

        i += 1

    finish_current()
    return {"sheets": sheets}


def _style_cell(cell, font: Font | None = None, fill: PatternFill | None = None, border: Border | None = None, alignment: Alignment | None = None) -> None:
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment


def _build_frontpage_sheet(
    wb: Workbook,
    title: str,
    report_date: str,
    workspace_root: str,
    code_findings: list[dict[str, Any]],
    vulns: list[dict[str, Any]],
    sbom: list[dict[str, Any]],
    opts: dict[str, Any],
    thin_border: Border,
    header_font: Font,
    header_fill: PatternFill,
    header_font_white: Font,
    section_fill: PatternFill,
    title_fill: PatternFill,
    title_font: Font,
    wrap_align: Alignment,
) -> None:
    ws_front = wb.active
    ws_front.title = "FRONTPAGE"
    ws_front.merge_cells("A1:B1")
    ws_front["A1"] = title
    _style_cell(
        ws_front["A1"],
        font=title_font,
        fill=title_fill,
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
    )
    ws_front.row_dimensions[1].height = 36
    ws_front["A2"] = "Generated"
    ws_front["B2"] = report_date
    ws_front["A2"].font = header_font
    ws_front["A3"] = "Workspace"
    ws_front["B3"] = workspace_root
    ws_front["A3"].font = header_font
    ws_front.column_dimensions["A"].width = 28
    ws_front.column_dimensions["B"].width = 52

    # Summary section
    r = 5
    ws_front.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws_front.cell(row=r, column=1, value="Summary").font = header_font
    ws_front.cell(row=r, column=1).fill = section_fill
    ws_front.cell(row=r, column=1).border = thin_border
    ws_front.cell(row=r, column=2).fill = section_fill
    ws_front.cell(row=r, column=2).border = thin_border
    r += 1
    summary_rows = [("SAST findings (code)", len(code_findings))]
    if opts.get("dependencies") is not False:
        summary_rows.append(("SCA vulnerabilities (dependencies)", len(vulns)))
        summary_rows.append(("SBOM components", len(sbom)))
    for label, value in summary_rows:
        ws_front.cell(row=r, column=1, value=label).border = thin_border
        ws_front.cell(row=r, column=2, value=value).border = thin_border
        r += 1
    r += 1

    # SAST by severity
    ws_front.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws_front.cell(row=r, column=1, value="SAST by severity").font = header_font
    ws_front.cell(row=r, column=1).fill = section_fill
    ws_front.cell(row=r, column=1).border = thin_border
    ws_front.cell(row=r, column=2).fill = section_fill
    ws_front.cell(row=r, column=2).border = thin_border
    r += 1
    ws_front.cell(row=r, column=1, value="Severity").font = header_font_white
    ws_front.cell(row=r, column=1).fill = header_fill
    ws_front.cell(row=r, column=1).border = thin_border
    ws_front.cell(row=r, column=2, value="Count").font = header_font_white
    ws_front.cell(row=r, column=2).fill = header_fill
    ws_front.cell(row=r, column=2).border = thin_border
    r += 1
    sast_sev = _severity_sast(code_findings)
    for sev, count in sorted(sast_sev.items()):
        ws_front.cell(row=r, column=1, value=sev).border = thin_border
        ws_front.cell(row=r, column=2, value=count).border = thin_border
        r += 1
    r += 1

    if opts.get("dependencies") is not False:
        # SCA by severity
        ws_front.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws_front.cell(row=r, column=1, value="SCA by severity").font = header_font
        ws_front.cell(row=r, column=1).fill = section_fill
        ws_front.cell(row=r, column=1).border = thin_border
        ws_front.cell(row=r, column=2).fill = section_fill
        ws_front.cell(row=r, column=2).border = thin_border
        r += 1
        ws_front.cell(row=r, column=1, value="Severity").font = header_font_white
        ws_front.cell(row=r, column=1).fill = header_fill
        ws_front.cell(row=r, column=1).border = thin_border
        ws_front.cell(row=r, column=2, value="Count").font = header_font_white
        ws_front.cell(row=r, column=2).fill = header_fill
        ws_front.cell(row=r, column=2).border = thin_border
        r += 1
        sca_sev = _severity_sca(vulns)
        for sev, count in sorted(sca_sev.items()):
            ws_front.cell(row=r, column=1, value=sev).border = thin_border
            ws_front.cell(row=r, column=2, value=count).border = thin_border
            r += 1


def _build_infra_sheet(
    wb: Workbook,
    images: list[dict[str, Any]],
    findings_infra: list[dict[str, Any]],
    thin_border: Border,
    header_font_white: Font,
    header_fill: PatternFill,
    wrap_align: Alignment,
) -> None:
    images = images or []
    findings_infra = findings_infra or []
    ws_infra = wb.create_sheet("INFRA")
    inf_headers_img = ["File", "Line", "Image", "Source"]
    for col, h in enumerate(inf_headers_img, start=1):
        cell = ws_infra.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = wrap_align
    if images:
        for row, im in enumerate(images, start=2):
            ws_infra.cell(row=row, column=1, value=_safe(im.get("file")))
            ws_infra.cell(row=row, column=2, value=im.get("line"))
            ws_infra.cell(row=row, column=3, value=_safe(im.get("image_ref") or im.get("image")))
            ws_infra.cell(row=row, column=4, value=_safe(im.get("source")))
            for c in range(1, 5):
                ws_infra.cell(row=row, column=c).alignment = wrap_align
    else:
        ws_infra.cell(row=2, column=1, value="No images")
        ws_infra.cell(row=2, column=1).alignment = wrap_align
    r_infra = 2 + max(len(images), 1)
    inf_headers_fin = ["Rule ID", "Severity", "File", "Line", "Message"]
    for col, h in enumerate(inf_headers_fin, start=1):
        cell = ws_infra.cell(row=r_infra, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = wrap_align
    r_infra += 1
    if findings_infra:
        for f in findings_infra:
            vulns = f.get("vulnerabilities") or []
            # For infra.image-vulnerability, expand each CVE into its own row instead of a summarized message
            if f.get("rule_id") == "infra.image-vulnerability" and vulns:
                base_file = _safe(f.get("file"))
                base_line = f.get("line")
                image_ref = _safe(f.get("image_ref") or f.get("image"))
                for v in vulns:
                    vid = _safe(v.get("vulnerability_id"))
                    pkg = _safe(v.get("pkg_name"))
                    sev = _safe(v.get("severity") or f.get("severity"))
                    title = _safe(v.get("title"))
                    desc = _safe(v.get("description"))
                    parts = []
                    if image_ref:
                        parts.append(f"Image {image_ref}")
                    if pkg:
                        parts.append(f"Package {pkg}")
                    if vid:
                        parts.append(f"Vulnerability {vid}")
                    if title:
                        parts.append(title)
                    elif desc:
                        parts.append(desc)
                    message = " – ".join([p for p in parts if p])
                    ws_infra.cell(row=r_infra, column=1, value=_safe(f.get("rule_id")))
                    ws_infra.cell(row=r_infra, column=2, value=sev)
                    ws_infra.cell(row=r_infra, column=3, value=base_file)
                    ws_infra.cell(row=r_infra, column=4, value=base_line)
                    ws_infra.cell(row=r_infra, column=5, value=message)
                    for c in range(1, 6):
                        ws_infra.cell(row=r_infra, column=c).alignment = wrap_align
                    r_infra += 1
            else:
                ws_infra.cell(row=r_infra, column=1, value=_safe(f.get("rule_id")))
                ws_infra.cell(row=r_infra, column=2, value=_safe(f.get("severity")))
                ws_infra.cell(row=r_infra, column=3, value=_safe(f.get("file")))
                ws_infra.cell(row=r_infra, column=4, value=f.get("line"))
                ws_infra.cell(row=r_infra, column=5, value=_safe(f.get("message")))
                for c in range(1, 6):
                    ws_infra.cell(row=r_infra, column=c).alignment = wrap_align
                r_infra += 1
    else:
        ws_infra.cell(row=r_infra, column=1, value="No findings")
        ws_infra.cell(row=r_infra, column=1).alignment = wrap_align
    ws_infra.column_dimensions["A"].width = 32
    ws_infra.column_dimensions["B"].width = 10
    ws_infra.column_dimensions["C"].width = 24
    ws_infra.column_dimensions["D"].width = 38
    ws_infra.column_dimensions["E"].width = 48


def _build_table_block(
    wb: Workbook,
    sheet_name: str,
    table_def: dict[str, Any],
    report: dict[str, Any],
    thin_border: Border,
    header_font_white: Font,
    header_fill: PatternFill,
    wrap_align: Alignment,
    severity_fills: dict[str, PatternFill] | None = None,
    ws=None,
    start_row: int = 1,
):
    """Construye una tabla en una hoja (nueva o existente) y devuelve (ws, next_row)."""
    source_path = table_def.get("source")
    if not source_path:
        return ws, start_row

    # Obtener filas desde el JSON canónico
    rows: list[dict[str, Any]] = []
    base = _get_path(report, source_path) or []
    if isinstance(base, list):
        rows = base

    # Expandir vulnerabilidades de findings infra cuando se solicite explícitamente
    expand_mode = (table_def.get("expand") or "").strip().lower()
    if expand_mode == "vulnerabilities" and source_path == "infrastructure.findings":
        expanded: list[dict[str, Any]] = []
        for f in rows:
            if not isinstance(f, dict):
                continue
            if f.get("rule_id") != "infra.image-vulnerability":
                continue
            base_file = _safe(f.get("file"))
            base_line = f.get("line")
            image_ref = _safe(f.get("image_ref") or f.get("image"))
            for v in f.get("vulnerabilities") or []:
                if not isinstance(v, dict):
                    continue
                expanded.append(
                    {
                        "image_ref": image_ref,
                        "file": base_file,
                        "line": base_line,
                        "vulnerability_id": _safe(v.get("vulnerability_id")),
                        "pkg_name": _safe(v.get("pkg_name")),
                        "severity": _safe(v.get("severity") or f.get("severity")),
                        "title_or_description": _safe(v.get("title")) or _safe(v.get("description")),
                    }
                )
        rows = expanded

    if not isinstance(rows, list) or not rows:
        return ws, start_row

    if ws is None:
        ws = wb.create_sheet(str(sheet_name))
        row_offset = 1
    else:
        row_offset = start_row

    columns = table_def.get("columns") or []

    # Cabecera
    for col_idx, col_def in enumerate(columns, start=1):
        header_text = str(col_def.get("header", ""))
        cell = ws.cell(row=row_offset, column=col_idx, value=header_text)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = wrap_align
        width = col_def.get("width")
        if isinstance(width, (int, float)) and width > 0:
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = float(width)

    # Filas de datos
    data_start = row_offset + 1
    current_row = data_start
    for item in rows:
        for col_idx, col_def in enumerate(columns, start=1):
            value_path = col_def.get("value")
            v = _get_path(item, str(value_path)) if value_path else None
            cell = ws.cell(row=current_row, column=col_idx, value=_safe(v))
            wrap = col_def.get("wrap")
            if wrap is None or wrap:
                cell.alignment = wrap_align
            number_format = col_def.get("number_format")
            if number_format:
                cell.number_format = str(number_format)
            style_name = (col_def.get("style") or "").strip().lower()
            if style_name == "severity" and severity_fills and v:
                sev_key = str(v).strip().lower()
                fill = severity_fills.get(sev_key)
                if fill:
                    cell.fill = fill
        current_row += 1

    return ws, current_row


def write_excel(
    report_dir: str,
    report: dict[str, Any],
    opts: dict[str, Any],
) -> list[tuple[str, int]]:
    """Write panorama-report.xlsx from canonical report.

    Sin plantilla XLSX, se mantiene el layout clásico:
    FRONTPAGE, SAST, SCA, INFRA (cuando infra está activado) y LICENSES.

    Con plantilla XLSX (explícita o la de `templates/panorama-xlsx-template.md`),
    todas las hojas/tablas (SAST, SCA, INFRA, LICENSES, etc.) se generan desde
    la plantilla usando rutas contra el JSON canónico. Solo FRONTPAGE sigue
    siendo programática por ahora. Returns [(path, size), ...].
    """
    wb = Workbook()
    title = opts.get("report_title") or report.get("metadata", {}).get("report_title") or "RootCause Panorama Report"
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    workspace_root = opts.get("workspace_root") or report.get("metadata", {}).get("workspace_root") or ""
    code_findings = report.get("sast", {}).get("findings", [])
    vulns = report.get("dependency_vulnerabilities", {}).get("vulnerabilities", [])
    sbom = report.get("sbom", {}).get("components", [])
    images = report.get("infrastructure", {}).get("images", [])
    findings_infra = report.get("infrastructure", {}).get("findings", [])

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # blue
    header_font_white = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")  # light gray
    title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")  # dark blue
    title_font = Font(bold=True, size=16, color="FFFFFF")
    wrap_align = Alignment(wrap_text=True, vertical="top")

    severity_fills = {
        "critical": PatternFill(start_color="C00000", end_color="C00000", fill_type="solid"),
        "high": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "medium": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        "low": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "info": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
    }

    # ---- FRONTPAGE (always programmatic) ----
    _build_frontpage_sheet(
        wb=wb,
        title=title,
        report_date=report_date,
        workspace_root=workspace_root,
        code_findings=code_findings,
        vulns=vulns,
        sbom=sbom,
        opts=opts,
        thin_border=thin_border,
        header_font=header_font,
        header_fill=header_fill,
        header_font_white=header_font_white,
        section_fill=section_fill,
        title_fill=title_fill,
        title_font=title_font,
        wrap_align=wrap_align,
    )

    # Resolve template path if configured or default template exists.
    template_def: dict[str, Any] | None = None
    template_path = opts.get("xlsx_template") or ""
    if template_path:
        if not os.path.isabs(template_path):
            plugin_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            template_path = os.path.join(plugin_dir, template_path)
        if os.path.isfile(template_path):
            try:
                template_def = _parse_xlsx_template(template_path)
            except Exception as exc:  # pragma: no cover - defensive
                log.warning("Failed to parse XLSX template '%s': %s", template_path, exc)
                template_def = None
        else:
            log.warning("XLSX template '%s' does not exist; falling back to built-in layout", template_path)
    else:
        plugin_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        default_tpl = os.path.join(plugin_dir, "templates", "panorama-xlsx-template.md")
        if os.path.isfile(default_tpl):
            try:
                template_def = _parse_xlsx_template(default_tpl)
            except Exception as exc:  # pragma: no cover - defensive
                log.warning("Failed to parse default XLSX template '%s': %s", default_tpl, exc)
                template_def = None

    # ---- SAST/SCA/INFRA/LICENSES vía plantilla cuando exista ----
    if template_def and template_def.get("sheets"):
        # Agrupar todas las tablas por nombre de hoja
        grouped: dict[str, list[dict[str, Any]]] = {}
        for sheet_def in template_def.get("sheets", []):
            sheet_type = (sheet_def.get("type") or "table").lower()
            if sheet_type != "table":
                continue
            raw_name = sheet_def.get("name") or "Sheet"
            sheet_name = str(raw_name)
            if sheet_name.upper() == "FRONTPAGE":
                # FRONTPAGE sigue siendo programática
                continue

            tables = sheet_def.get("tables") or []
            # Compatibilidad: si hay columnas pero no tables, construir tabla única
            if not tables and sheet_def.get("columns"):
                tables = [{"source": sheet_def.get("source"), "columns": sheet_def["columns"]}]

            if not tables:
                continue

            grouped.setdefault(sheet_name, []).extend(tables)

        for sheet_name, tables in grouped.items():
            # Respetar flags de activación de datos
            upper = sheet_name.upper()
            if upper == "SAST" and not code_findings:
                continue
            if upper == "SCA" and (opts.get("dependencies") is False or not vulns):
                continue
            if upper.startswith("INFRA") and (opts.get("infra") is False):
                continue
            if upper == "LICENSES" and (opts.get("licenses") is False or not sbom):
                continue

            ws_sheet = None
            row_cursor = 1
            for table_def in tables:
                ws_sheet, row_cursor = _build_table_block(
                    wb=wb,
                    sheet_name=sheet_name,
                    table_def=table_def,
                    report=report,
                    thin_border=thin_border,
                    header_font_white=header_font_white,
                    header_fill=header_fill,
                    wrap_align=wrap_align,
                    severity_fills=severity_fills,
                    ws=ws_sheet,
                    start_row=row_cursor,
                )
                row_cursor += 1  # una fila en blanco entre bloques
    else:
        # ---- Legacy SAST/SCA/LICENSES layout ----
        ws_sast = wb.create_sheet("SAST", 1)
        sast_headers = [
            "Finding ID",
            "Rule ID",
            "File",
            "Line",
            "Column",
            "Severity",
            "Message",
            "Excerpt",
            "Remediation",
            "Context",
        ]
        for col, h in enumerate(sast_headers, start=1):
            cell = ws_sast.cell(row=1, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = wrap_align
        ws_sast.column_dimensions["A"].width = 12
        ws_sast.column_dimensions["B"].width = 22
        ws_sast.column_dimensions["C"].width = 38
        ws_sast.column_dimensions["D"].width = 8
        ws_sast.column_dimensions["E"].width = 8
        ws_sast.column_dimensions["F"].width = 10
        ws_sast.column_dimensions["G"].width = 42
        ws_sast.column_dimensions["H"].width = 32
        ws_sast.column_dimensions["I"].width = 32
        ws_sast.column_dimensions["J"].width = 28

        for row, f in enumerate(code_findings, start=2):
            ws_sast.cell(row=row, column=1, value=_safe(f.get("id")))
            ws_sast.cell(row=row, column=2, value=_safe(f.get("rule_id")))
            ws_sast.cell(row=row, column=3, value=_safe(f.get("file") or f.get("path")))
            ws_sast.cell(row=row, column=4, value=f.get("line"))
            ws_sast.cell(row=row, column=5, value=f.get("column"))
            ws_sast.cell(row=row, column=6, value=_safe(f.get("severity")))
            ws_sast.cell(row=row, column=7, value=_safe(f.get("message")))
            ws_sast.cell(row=row, column=8, value=_safe(f.get("excerpt")))
            ws_sast.cell(row=row, column=9, value=_safe(f.get("remediation")))
            ws_sast.cell(row=row, column=10, value=_safe(f.get("context")))
            for c in range(1, 11):
                ws_sast.cell(row=row, column=c).alignment = wrap_align

        sheet_idx = 2
        if opts.get("dependencies") is not False:
            ws_sca = wb.create_sheet("SCA", sheet_idx)
            sheet_idx += 1
            sca_headers = [
                "Vuln ID",
                "Package",
                "Version",
                "Ecosystem",
                "File",
                "Line",
                "Severity",
                "Description",
                "Fixed In",
                "Published",
                "Modified",
                "References",
            ]
            for col, h in enumerate(sca_headers, start=1):
                cell = ws_sca.cell(row=1, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = wrap_align
            for row, v in enumerate(vulns, start=2):
                sev = _sca_severity_display(
                    v.get("severity") or severity_for_vuln(_safe(v.get("vuln_id", "")))
                )
                ws_sca.cell(row=row, column=1, value=_safe(v.get("vuln_id")))
                ws_sca.cell(row=row, column=2, value=_safe(v.get("name")))
                ws_sca.cell(row=row, column=3, value=_safe(v.get("version")))
                ws_sca.cell(row=row, column=4, value=_safe(v.get("ecosystem")))
                ws_sca.cell(row=row, column=5, value=_safe(v.get("file")))
                ws_sca.cell(row=row, column=6, value=v.get("line"))
                ws_sca.cell(row=row, column=7, value=sev)
                ws_sca.cell(row=row, column=8, value=_safe(v.get("description")))
                ws_sca.cell(row=row, column=9, value=_safe(v.get("fixed_in")))
                ws_sca.cell(row=row, column=10, value=_safe(v.get("published")))
                ws_sca.cell(row=row, column=11, value=_safe(v.get("modified")))
                ws_sca.cell(row=row, column=12, value=_safe(v.get("references")))
                for c in range(1, 13):
                    ws_sca.cell(row=row, column=c).alignment = wrap_align
            ws_sca.column_dimensions["A"].width = 22
            ws_sca.column_dimensions["B"].width = 10
            ws_sca.column_dimensions["C"].width = 24
            ws_sca.column_dimensions["D"].width = 14
            ws_sca.column_dimensions["E"].width = 12
            ws_sca.column_dimensions["F"].width = 32
            ws_sca.column_dimensions["G"].width = 8
            ws_sca.column_dimensions["H"].width = 48
            ws_sca.column_dimensions["I"].width = 14
            ws_sca.column_dimensions["J"].width = 12
            ws_sca.column_dimensions["K"].width = 12
            ws_sca.column_dimensions["L"].width = 28

        if opts.get("licenses") is not False:
            ws_lic = wb.create_sheet("LICENSES", sheet_idx)
            lic_headers = [
                "PURL",
                "Component Name",
                "Version",
                "Ecosystem",
                "Line",
                "Type",
                "License",
            ]
            for col, h in enumerate(lic_headers, start=1):
                cell = ws_lic.cell(row=1, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = wrap_align
            for row, c in enumerate(sbom, start=2):
                name = _safe(c.get("name"))
                version = _safe(c.get("version"))
                eco = _safe(c.get("ecosystem"))
                ws_lic.cell(row=row, column=1, value=_purl(eco, name, version))
                ws_lic.cell(row=row, column=2, value=name)
                ws_lic.cell(row=row, column=3, value=version)
                ws_lic.cell(row=row, column=4, value=eco)
                ws_lic.cell(row=row, column=5, value=c.get("line"))
                ws_lic.cell(row=row, column=6, value=_safe(c.get("type")) or "library")
                ws_lic.cell(row=row, column=7, value=_safe(c.get("license")) or "N/A")
                for col in range(1, 8):
                    ws_lic.cell(row=row, column=col).alignment = wrap_align
            ws_lic.column_dimensions["A"].width = 48
            ws_lic.column_dimensions["B"].width = 28
            ws_lic.column_dimensions["C"].width = 14
            ws_lic.column_dimensions["D"].width = 12
            ws_lic.column_dimensions["E"].width = 8
            ws_lic.column_dimensions["F"].width = 10
            ws_lic.column_dimensions["G"].width = 14

    # ---- INFRA legacy sin plantilla ----
    # Solo cuando no hay plantilla XLSX se mantiene la hoja INFRA programática.
    if not (template_def and template_def.get("sheets")):
        if opts.get("infra") is not False:
            _build_infra_sheet(
                wb=wb,
                images=images,
                findings_infra=findings_infra,
                thin_border=thin_border,
                header_font_white=header_font_white,
                header_fill=header_fill,
                wrap_align=wrap_align,
            )

    path = os.path.join(report_dir, "panorama-report.xlsx")
    wb.save(path)
    return [(path, os.path.getsize(path))]
